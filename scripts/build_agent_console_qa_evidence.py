from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT = Path(__file__).resolve().parents[1] / "artifacts" / "Agent-Console-QA-Evidence.xlsx"

rows = [
    (1, "English prospect handoff", "EN", "Maria trusted iPhone", "External device QA"),
    (2, "Transferencia de cliente", "ES", "Maria trusted iPhone", "External device QA"),
    (3, "Simultaneous claim protection", "EN/ES", "Maria iPhone + Carlos browser", "Two-device QA"),
    (4, "Dispositivo no confiable", "ES", "Untrusted browser", "Negative authorization QA"),
    (5, "Outbound message fails closed", "EN", "Controlled API test", "Backend QA"),
    (6, "Proteccion de conversacion real", "ES", "Controlled API test", "Backend QA"),
    (7, "Availability routing", "EN/ES", "Local automated validation", "Code QA"),
    (8, "Sesion y cierre seguro", "ES", "Maria trusted iPhone", "Session QA"),
]

wb = Workbook()
ws = wb.active
ws.title = "QA Evidence"

headers = [
    "ID", "Scenario", "Language", "Required environment", "Test class",
    "Status", "Tester", "Date/time", "Device/browser", "Conversation ID",
    "Expected result", "Actual result", "Evidence link", "Incident/fix", "Retest date"
]
ws.append(headers)
for row in rows:
    ws.append((*row, "PENDING", "", "", "", "", "See EN-ES scenario document", "", "", "", ""))

header_fill = PatternFill("solid", fgColor="111827")
header_font = Font(color="FFFFFF", bold=True)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

status_validation = DataValidation(type="list", formula1='"PENDING,PASS,FAIL,BLOCKED"', allow_blank=False)
ws.add_data_validation(status_validation)
status_validation.add("F2:F200")

fills = {
    "PASS": "C6EFCE",
    "FAIL": "FFC7CE",
    "BLOCKED": "FFEB9C",
    "PENDING": "D9EAF7",
}
for status, color in fills.items():
    ws.conditional_formatting.add(
        "F2:F200",
        FormulaRule(formula=[f'$F2="{status}"'], fill=PatternFill("solid", fgColor=color)),
    )

widths = {
    "A": 8, "B": 34, "C": 12, "D": 30, "E": 25, "F": 14, "G": 18,
    "H": 22, "I": 28, "J": 22, "K": 34, "L": 34, "M": 34, "N": 34, "O": 18,
}
for column, width in widths.items():
    ws.column_dimensions[column].width = width
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:O{ws.max_row}"

summary = wb.create_sheet("Summary")
summary.append(["Agent Console QA Release Gate", "Value"])
summary.append(["Total scenarios", "=COUNTA('QA Evidence'!A2:A200)"])
summary.append(["PASS", '=COUNTIF(\'QA Evidence\'!F2:F200,"PASS")'])
summary.append(["FAIL", '=COUNTIF(\'QA Evidence\'!F2:F200,"FAIL")'])
summary.append(["BLOCKED", '=COUNTIF(\'QA Evidence\'!F2:F200,"BLOCKED")'])
summary.append(["PENDING", '=COUNTIF(\'QA Evidence\'!F2:F200,"PENDING")'])
summary.append(["Release decision", '=IF(B3=B2,"READY FOR REVIEW","HOLD")'])
summary.append(["Safety rule", "Live messages, real conversations and push remain disabled until every scenario is PASS."])
for cell in summary[1]:
    cell.fill = header_fill
    cell.font = header_font
summary.column_dimensions["A"].width = 32
summary.column_dimensions["B"].width = 85
summary.freeze_panes = "A2"
for row in summary.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT)
print(OUTPUT)
